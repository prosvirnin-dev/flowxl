"""Random-access sheets: cells can be written in any order."""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from enum import Enum
from typing import TYPE_CHECKING, Final, Self

from flowxl.api.base_sheet import BaseSheet
from flowxl.api.settings import WorkbookSettings
from flowxl.backend.openpyxl_backend import OpenpyxlRandomAccessSheet
from flowxl.integrations.protocol import TabularSource
from flowxl.primitives.cell import CellRange, CellRef, letter_to_index
from flowxl.primitives.exceptions import UsageError
from flowxl.primitives.sheet_name import SheetName
from flowxl.primitives.types import infer_logical_type
from flowxl.styles.spec import StyleSpec

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

__all__ = ['KEEP', 'Sheet']


class _Sentinel(Enum):
    """Single-member enum used as a typing-friendly sentinel."""

    KEEP = 'keep'


KEEP: Final[_Sentinel] = _Sentinel.KEEP
"""Marker meaning "do not change the value". None is a real cell value that clears the cell."""

# KEEP нельзя заменить на None.
# В Excel запись None очищает ячейку, а нам нужен отдельный сигнал «стиль поменять, значение оставить».


class Sheet(BaseSheet):
    """A sheet whose cells can be written in any order.

    Methods that a streaming sheet does not have live here: ``cell``, ``read``,
    ``merge``, ``freeze``. The shared ``frame`` algorithm lives on :class:`BaseSheet`.
    """

    def __init__(
        self,
        name: SheetName,
        backend: OpenpyxlRandomAccessSheet,
        settings: WorkbookSettings,
    ) -> None:
        """Bind a validated name to a random-access sheet backend.

        Args:
            name: Validated sheet name.
            backend: openpyxl adapter for this sheet.
            settings: Workbook-wide configuration.
        """
        super().__init__(name, backend, settings)
        # Уточняем тип бэкенда: родителю достаточно общего, здесь нужен полный.
        self._backend: OpenpyxlRandomAccessSheet = backend

    def cell(
        self,
        ref: CellRef,
        value: object | _Sentinel = KEEP,
        style: StyleSpec | None = None,
    ) -> Self:
        """Write a value and optionally a style to one cell.

        Args:
            ref: Target cell.
            value: New value. Pass nothing to keep the current one. ``None``
                is a real value that clears the cell.
            style: Style to apply. Defaults to the theme style for the value's
                logical type.

        Returns:
            This sheet, for chaining.
        """
        if value is KEEP:
            spec: StyleSpec = style or self._settings.theme.default_style()
            self._backend.apply_style(ref, spec)
            self._extend_bounds(CellRange(ref, ref))
            return self
        # Если стиль не передали, тема выбирает оформление по смыслу значения.
        spec = style or self._settings.theme.style_for(infer_logical_type(value))
        return self._put(ref, value, spec)

    def row(self, at: CellRef, values: Sequence[object], style: StyleSpec | None = None) -> Self:
        """Write values left to right starting at ``at``.

        Args:
            at: First cell of the row.
            values: Values to write.
            style: Optional style applied to every cell.

        Returns:
            This sheet, for chaining.
        """
        for offset, value in enumerate(values):
            self.cell(at.offset(cols=offset), value, style)
        return self

    def column(self, at: CellRef, values: Sequence[object], style: StyleSpec | None = None) -> Self:
        """Write values top to bottom starting at ``at``.

        Args:
            at: First cell of the column.
            values: Values to write.
            style: Optional style applied to every cell.

        Returns:
            This sheet, for chaining.
        """
        for offset, value in enumerate(values):
            self.cell(at.offset(rows=offset), value, style)
        return self

    def formula(self, ref: CellRef, expression: str, style: StyleSpec | None = None) -> Self:
        """Write an Excel formula.

        Russian names and separators are translated to the en-US form the
        file stores: ``=СУММ(A1;A2)`` becomes ``=SUM(A1,A2)``, and ``1,5``
        becomes ``1.5``. An English formula is stored as written. This
        library never evaluates the result.

        Args:
            ref: Target cell.
            expression: Formula text, starting with ``=``.
            style: Optional style. Defaults to the theme default.

        Returns:
            This sheet, for chaining.

        Raises:
            UsageError: If ``expression`` does not start with ``=``, or a
                Cyrillic function name is unknown.
        """
        if not expression.startswith('='):
            raise UsageError(
                f'A formula must start with "=", got {expression!r}. '
                f'Write it as "={expression}" if that was the intent.'
            )
        spec: StyleSpec = style or self._settings.theme.default_style()
        return self._put(ref, expression, spec)

    def read(self, ref: CellRef) -> object:
        """Return a cell's current value.

        Args:
            ref: Cell to read.

        Returns:
            The stored value, possibly ``None``.
        """
        return self._backend.read_cell(ref)

    def freeze(self, ref: CellRef) -> Self:
        """Freeze everything above and left of ``ref``.

        Args:
            ref: First unfrozen cell.

        Returns:
            This sheet, for chaining.
        """
        self._backend.set_freeze(ref)
        return self

    def pin_header(self, rows: int = 1) -> Self:
        """Freeze the top rows.

        Args:
            rows: How many rows to pin. At least 1.

        Returns:
            This sheet, for chaining.

        Raises:
            UsageError: If ``rows`` is below 1.
        """
        if rows < 1:
            raise UsageError(f'pin_header needs at least one row, got {rows}.')
        return self.freeze(CellRef(row=rows + 1, col=1))

    def merge(self, area: CellRange) -> Self:
        """Merge a rectangular area into a single cell.

        Args:
            area: Cells to merge.

        Returns:
            This sheet, for chaining.
        """
        self._backend.merge(area)
        self._extend_bounds(area)
        return self

    def widths(self, values: Mapping[int | str, float]) -> Self:
        """Set column widths.

        Args:
            values: Width per column, keyed by 1-based index or by letter.

        Returns:
            This sheet, for chaining.
        """
        for key, width in values.items():
            col: int = key if isinstance(key, int) else letter_to_index(key)
            self._backend.set_column_width(col, width)
        return self

    def heights(self, values: Mapping[int, float]) -> Self:
        """Set row heights.

        Args:
            values: Height per row, keyed by 1-based row index.

        Returns:
            This sheet, for chaining.
        """
        for row, height in values.items():
            self._backend.set_row_height(row, height)
        return self

    def autofit(self, *, max_width: float | None = None, only: Iterable[int] | None = None) -> Self:
        """Size columns to the widest value written through this library.

        Args:
            max_width: Upper cap. Defaults to the workbook setting.
            only: Restrict to these 1-based column indices.

        Returns:
            This sheet, for chaining.
        """
        cap: float = max_width or self._settings.autofit_max_width
        padding: float = self._settings.autofit_padding
        wanted: set[int] | None = set(only) if only is not None else None
        for col, length in self._content_width.items():
            if wanted is not None and col not in wanted:
                continue
            self._backend.set_column_width(col, min(length + padding, cap))
        return self

    def raw_openpyxl(self) -> Worksheet:
        """Return the underlying openpyxl worksheet.

        After this call, :meth:`~flowxl.api.base_sheet.BaseSheet.autofilter`
        will not guess bounds.

        Returns:
            The raw worksheet.
        """
        self._bounds_trusted = False
        return self._backend.worksheet

    def _put(self, ref: CellRef, value: object, style: StyleSpec) -> Self:
        """Write a value through the single funnel used by every write.

        Args:
            ref: Target cell.
            value: Value to store.
            style: Style to apply.

        Returns:
            This sheet, for chaining.
        """
        # Одна воронка: запись, ширина для autofit и границы для autofilter.
        self._backend.write_cell(ref, value, style)
        self._track_width(ref.col, value)
        self._extend_bounds(CellRange(ref, ref))
        return self

    def _write_header(self, at: CellRef, columns: Sequence[str]) -> None:
        """Write the header one cell at a time.

        Args:
            at: First header cell.
            columns: Column names.
        """
        header_style: StyleSpec = self._settings.theme.header_style()
        for offset, column in enumerate(columns):
            self._put(at.offset(cols=offset), column, header_style)

    def _write_body(
        self,
        at: CellRef,
        source: TabularSource,
        styles: Sequence[StyleSpec],
    ) -> int:
        """Write data rows one cell at a time.

        Args:
            at: First data cell.
            source: Tabular data.
            styles: One style per column.

        Returns:
            Number of data rows written.
        """
        written: int = 0
        for index, values in enumerate(source.iter_rows()):
            row_at: CellRef = at.offset(rows=index)
            for offset, (value, style) in enumerate(zip(values, styles, strict=True)):
                self._put(row_at.offset(cols=offset), value, style)
            written += 1
        return written
