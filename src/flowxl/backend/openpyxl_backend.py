"""The only module that talks to sheets and workbooks through openpyxl.

The rest of the library speaks :class:`~flowxl.primitives.cell.CellRef` and
:class:`~flowxl.styles.spec.StyleSpec`. This file translates those types
into openpyxl calls.
"""

from __future__ import annotations

from collections.abc import Sequence
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook as OpxWorkbook
from openpyxl import load_workbook
from openpyxl.cell import Cell, WriteOnlyCell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet as OpxWorksheet

from flowxl.backend.registry import AppliedStyle, StyleRegistry
from flowxl.primitives.cell import CellRange, CellRef, index_to_letter
from flowxl.primitives.exceptions import CorruptedWorkbookError, UsageError, WorkbookOpenError, WorkbookSaveError
from flowxl.primitives.formula import to_excel_formula
from flowxl.primitives.sheet_name import SheetName
from flowxl.styles.spec import StyleSpec

__all__ = [
    'OpenpyxlRandomAccessSheet',
    'OpenpyxlStreamingSheet',
    'OpenpyxlStreamingWorkbook',
    'OpenpyxlWorkbook',
]


class OpenpyxlSheetBase:
    """Shared operations: column width, row height, freeze and filter.

    Both a regular sheet and a streaming sheet need these. Cell writes differ,
    so they live on the subclasses.
    """

    def __init__(self, worksheet: OpxWorksheet, registry: StyleRegistry) -> None:
        """Wrap one openpyxl worksheet and the workbook's style registry.

        Args:
            worksheet: An openpyxl worksheet.
            registry: Style registry of the parent workbook.
        """
        self._ws: OpxWorksheet = worksheet
        self._registry: StyleRegistry = registry

    @property
    def worksheet(self) -> OpxWorksheet:
        """Return the raw worksheet.

        Only :meth:`~flowxl.api.sheet.Sheet.raw_openpyxl` should use this.

        Returns:
            The openpyxl worksheet.
        """
        return self._ws

    def set_column_width(self, col: int, width: float) -> None:
        """Set the width of one column.

        Args:
            col: Column index counted from 1.
            width: Width in Excel character units.
        """
        self._ws.column_dimensions[index_to_letter(col)].width = width

    def set_row_height(self, row: int, height: float) -> None:
        """Set the height of one row.

        Args:
            row: Row index counted from 1.
            height: Height in points.
        """
        self._ws.row_dimensions[row].height = height

    def set_freeze(self, ref: CellRef | None) -> None:
        """Freeze panes at ``ref``, or clear freeze if ``ref`` is None.

        Args:
            ref: First unfrozen cell, or ``None`` to clear freeze.
        """
        self._ws.freeze_panes = ref.to_a1() if ref else None

    def set_autofilter(self, area: CellRange | None) -> None:
        """Enable the autofilter over ``area``, or clear it if ``area`` is None.

        Args:
            area: Filter range, or ``None`` to remove the filter.
        """
        self._ws.auto_filter.ref = area.to_a1() if area else None

    def protect_sheet(self, password: str | None = None) -> None:
        """Turn on sheet protection.

        Args:
            password: Optional password. Omit it to lock the sheet without one.
        """
        if password:
            self._ws.protection.set_password(password)
            return
        self._ws.protection.enable()

    def unprotect_sheet(self) -> None:
        """Turn off sheet protection."""
        self._ws.protection.disable()

    def add_dropdown(self, area: CellRange, source: str, *, allow_blank: bool) -> None:
        """Restrict ``area`` to a list of values.

        Args:
            area: Cells that show the dropdown.
            source: ``'"a,b,c"'`` or an absolute range such as ``$A$1:$A$3``.
            allow_blank: Whether an empty cell is accepted.
        """
        sqref: str = area.start.to_a1() if area.start == area.end else area.to_a1()
        validation: DataValidation = DataValidation(
            type='list',
            formula1=source,
            allow_blank=allow_blank,
            showDropDown=False,
            showErrorMessage=True,
        )
        validation.add(sqref)
        self._ws.data_validations.append(validation)

    def _paint(self, cell: Cell | WriteOnlyCell, style: StyleSpec) -> None:
        """Apply a resolved style to an openpyxl cell.

        Args:
            cell: A regular cell or a write-only cell.
            style: Drawing to resolve and apply.
        """
        applied: AppliedStyle = self._registry.resolve(style)
        cell.style = applied.named_style
        if applied.number_format is not None:
            cell.number_format = applied.number_format


def _excel_value(value: object) -> object:
    """Normalize a formula string to the en-US form Excel stores.

    Other values pass through. Conversion lives here so every write path
    (cell, formula, frame, streaming append) gets it.

    Args:
        value: Value about to be stored.

    Returns:
        The same value, or a translated formula.
    """
    if isinstance(value, str) and value.startswith('='):
        return to_excel_formula(value)
    return value


class OpenpyxlRandomAccessSheet(OpenpyxlSheetBase):
    """Backend for a sheet whose cells can be written in any order."""

    def write_cell(self, ref: CellRef, value: object, style: StyleSpec) -> None:
        """Write a value and apply a named style.

        Args:
            ref: Target cell.
            value: Value to store.
            style: Style to register and apply.
        """
        cell = self._ws.cell(row=ref.row, column=ref.col, value=_excel_value(value))
        self._paint(cell, style)

    def apply_style(self, ref: CellRef, style: StyleSpec) -> None:
        """Apply a style without changing the value.

        Args:
            ref: Target cell.
            style: Style to register and apply.
        """
        self._paint(self._ws.cell(row=ref.row, column=ref.col), style)

    def read_cell(self, ref: CellRef) -> object:
        """Return a cell's current value.

        Args:
            ref: Cell to read.

        Returns:
            The stored value, possibly ``None``.
        """
        return self._ws.cell(row=ref.row, column=ref.col).value

    def merge(self, area: CellRange) -> None:
        """Merge a rectangle into a single cell.

        Args:
            area: Cells to merge.
        """
        self._ws.merge_cells(area.to_a1())


class OpenpyxlStreamingSheet(OpenpyxlSheetBase):
    """Backend for a sheet that only accepts rows from top to bottom."""

    def append_row(self, values: Sequence[object], styles: Sequence[StyleSpec]) -> None:
        """Append one row of styled cells.

        Args:
            values: Values left to right.
            styles: One style per value. Lengths must match.
        """
        cells: list[WriteOnlyCell] = []
        for value, spec in zip(values, styles, strict=True):
            cell: WriteOnlyCell = WriteOnlyCell(self._ws, value=_excel_value(value))
            self._paint(cell, spec)
            cells.append(cell)
        self._ws.append(cells)


class OpenpyxlWorkbook:
    """Workbook backend for creating and editing documents."""

    def __init__(self, workbook: OpxWorkbook) -> None:
        """Wrap an openpyxl workbook and give it its own style registry.

        Args:
            workbook: An openpyxl workbook.
        """
        self._wb: OpxWorkbook = workbook
        self._registry: StyleRegistry = StyleRegistry(workbook)

    @classmethod
    def create(cls) -> OpenpyxlWorkbook:
        """Create an empty workbook with no sheets.

        openpyxl always starts with a default sheet. We delete it so a new
        FlowXL workbook is truly empty until the caller asks for a sheet.

        Returns:
            A new backend.
        """
        workbook: OpxWorkbook = OpxWorkbook()
        default = workbook.active
        if default is not None:
            workbook.remove(default)
        return cls(workbook)

    @classmethod
    def load(cls, source: str | Path | bytes | BinaryIO, *, formulas: bool) -> OpenpyxlWorkbook:
        """Open an existing document.

        Args:
            source: A path, raw bytes or a binary stream.
            formulas: ``True`` keeps formulas as written. ``False`` reads the
                cache of computed values.

        Returns:
            A backend around the loaded workbook.

        Raises:
            CorruptedWorkbookError: If the file is not a readable xlsx.
            WorkbookOpenError: If the file cannot be read.
        """
        handle: str | Path | BinaryIO = BytesIO(source) if isinstance(source, bytes) else source
        try:
            # Флаг data_only=True читает кэш значений. Наш флаг formulas говорит наоборот.
            workbook: OpxWorkbook = load_workbook(handle, data_only=not formulas)
        except InvalidFileException as exc:
            raise CorruptedWorkbookError(f'{source!r} is not a readable xlsx document.') from exc
        except (OSError, KeyError) as exc:
            raise WorkbookOpenError(f'Could not open {source!r}: {exc}') from exc
        return cls(workbook)

    def has_sheet(self, name: SheetName) -> bool:
        """Return whether a sheet with this name exists.

        Args:
            name: Validated sheet name.

        Returns:
            True if the sheet is in the workbook.
        """
        return name.value in self._wb.sheetnames

    def get_sheet(self, name: SheetName) -> OpenpyxlRandomAccessSheet:
        """Return an existing sheet.

        Args:
            name: Validated sheet name.

        Returns:
            A random-access sheet backend.
        """
        return OpenpyxlRandomAccessSheet(self._wb[name.value], self._registry)

    def create_sheet(self, name: SheetName) -> OpenpyxlRandomAccessSheet:
        """Create a new sheet.

        Args:
            name: Validated sheet name.

        Returns:
            A random-access sheet backend.
        """
        worksheet: OpxWorksheet = self._wb.create_sheet(title=name.value)
        return OpenpyxlRandomAccessSheet(worksheet, self._registry)

    def remove_sheet(self, name: SheetName) -> None:
        """Delete a sheet.

        Args:
            name: Validated sheet name.

        Raises:
            UsageError: If no such sheet exists.
        """
        if name.value not in self._wb.sheetnames:
            raise UsageError(f'Sheet {name.value!r} does not exist.')
        self._wb.remove(self._wb[name.value])

    def sheet_names(self) -> tuple[SheetName, ...]:
        """Return the names of all sheets, in document order.

        Returns:
            Validated sheet names.
        """
        return tuple(SheetName(title) for title in self._wb.sheetnames)

    def save(self, target: str | Path | BinaryIO) -> None:
        """Write the workbook to a path or stream.

        Args:
            target: Destination path or binary stream.

        Raises:
            WorkbookSaveError: If the write failed.
        """
        try:
            self._wb.save(target)
        except OSError as exc:
            raise WorkbookSaveError(f'Could not write the workbook to {target!r}: {exc}') from exc

    def close(self) -> None:
        """Release resources held by openpyxl."""
        self._wb.close()


class OpenpyxlStreamingWorkbook:
    """Workbook backend for one-shot streaming writes."""

    def __init__(self, workbook: OpxWorkbook) -> None:
        """Wrap an openpyxl workbook in write-only mode.

        Args:
            workbook: An openpyxl workbook created with ``write_only=True``.
        """
        self._wb: OpxWorkbook = workbook
        self._registry: StyleRegistry = StyleRegistry(workbook)

    @classmethod
    def create(cls) -> OpenpyxlStreamingWorkbook:
        """Create a write-only workbook.

        Returns:
            A new streaming backend.
        """
        return cls(OpxWorkbook(write_only=True))

    def create_sheet(self, name: SheetName) -> OpenpyxlStreamingSheet:
        """Create a streaming sheet.

        Args:
            name: Validated sheet name.

        Returns:
            A streaming sheet backend.
        """
        worksheet = self._wb.create_sheet(title=name.value)
        return OpenpyxlStreamingSheet(worksheet, self._registry)

    def save(self, target: str | Path | BinaryIO) -> None:
        """Write the document once.

        A write-only workbook cannot be serialised twice. That is an openpyxl
        limit, and this method does not hide it.

        Args:
            target: Destination path or binary stream.

        Raises:
            WorkbookSaveError: If the write failed.
        """
        try:
            self._wb.save(target)
        except OSError as exc:
            raise WorkbookSaveError(f'Could not write the workbook to {target!r}: {exc}') from exc

    def close(self) -> None:
        """Release resources held by openpyxl."""
        self._wb.close()
